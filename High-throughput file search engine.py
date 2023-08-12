import tkinter as tk
from tkinter import filedialog, messagebox
import os
import shutil
from openpyxl import load_workbook

def main(raw_data_path,string_path,copy_to_path,recurisively):
    strings = []
    wb = load_workbook(filename=string_path)
    sheet = wb[wb.sheetnames[0]]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value != None:
                strings.append(cell.value)
    strings = set(strings)

    def copy_files(original_path,copy_to_path):            
        files_or_subfolders = os.listdir(original_path)
        for file in files_or_subfolders:
            copy_or_not = False
            for string in strings:
                if string.upper() in file.upper():
                    copy_or_not = True
            if copy_or_not:
                if os.path.exists(copy_to_path+'/'+file):
                    if os.path.isfile(copy_to_path+'/'+file):
                        os.remove(copy_to_path+'/'+file)
                    elif os.path.isdir(copy_to_path+'/'+file):
                        shutil.rmtree(copy_to_path+'/'+file)
                if os.path.isfile(original_path+'/'+file):
                    shutil.copy(original_path+'/'+file,copy_to_path)
                elif os.path.isdir(original_path+'/'+file):
                    shutil.copytree(original_path+'/'+file,copy_to_path+'/'+file)

    def copy_files_recursively(original_path,copy_to_path):            
        files_or_subfolders = os.listdir(original_path)
        for file in files_or_subfolders:
            copy_or_not = False
            for string in strings:
                if string.upper() in file.upper():
                    copy_or_not = True
            if copy_or_not:
                if os.path.exists(copy_to_path+'/'+file):
                    if os.path.isfile(copy_to_path+'/'+file):
                        os.remove(copy_to_path+'/'+file)
                    elif os.path.isdir(copy_to_path+'/'+file):
                        shutil.rmtree(copy_to_path+'/'+file)
                if os.path.isfile(original_path+'/'+file):
                    shutil.copy(original_path+'/'+file,copy_to_path)
                elif os.path.isdir(original_path+'/'+file):
                    shutil.copytree(original_path+'/'+file,copy_to_path+'/'+file)
            else:
                if os.path.isdir(original_path+'/'+file):
                    copy_files_recursively(original_path+'/'+file,copy_to_path)
                else:
                    continue
    
    if recurisively:
        copy_files_recursively(raw_data_path,copy_to_path)
    else:
        copy_files(raw_data_path,copy_to_path)

def select_raw_data_folder():
    raw_data_folder.set(filedialog.askdirectory())
    
def open_file():
    file_path = filedialog.askopenfilename()
    file_path_entry.delete(0, tk.END)
    file_path_entry.insert(0, file_path)

def select_copy_to_folder():
    copy_to_folder.set(filedialog.askdirectory())

def run_code():
    raw_data_path = raw_data_folder.get()
    string_path = file_path_entry.get()
    copy_to_path = copy_to_folder.get()
    recurisively = recurisively_check.get()
    main(raw_data_path,string_path,copy_to_path,recurisively)
    messagebox.showinfo("High-throughput File Search Engine", "Files/folders containing the strings in their names have been copied to the destination folder")
    
root = tk.Tk()
root.title("High-throughput File Search Engine")

raw_data_folder = tk.StringVar()
copy_to_folder = tk.StringVar()

raw_data_label = tk.Label(root, text="Original folder you hope to copy files/subfolders from:")
raw_data_label.grid(row=0, column=0)
raw_data_entry = tk.Entry(root, width=120, textvariable=raw_data_folder)
raw_data_entry.grid(row=0, column=1)
raw_data_button = tk.Button(root, text="Browse", command=select_raw_data_folder)
raw_data_button.grid(row=0, column=2)
raw_data_note = tk.Label(root, text='This app will copy all files/subfolders contain the strings imported below in their names (case-insensitive).')
raw_data_note.grid(row=1, column=1)

empty = tk.Label(root, text="")
empty.grid(row=2, column=1)
empty = tk.Label(root, text="")
empty.grid(row=3, column=1)

file_path_label = tk.Label(root, text="Excel file (xlsx) for strings:")
file_path_label.grid(row=4, column=0)
file_path_entry = tk.Entry(root, width=120)
file_path_entry.grid(row=4, column=1)
file_path_button = tk.Button(root, text="Select File", command=open_file)
file_path_button.grid(row=4, column=2)
file_path_note = tk.Label(root, text='Each cell in the Excel spreadsheet should contain only one string to be searched (case-insensitive).')
file_path_note.grid(row=5, column=1)

empty = tk.Label(root, text="")
empty.grid(row=6, column=1)
empty = tk.Label(root, text="")
empty.grid(row=7, column=1)

copy_to_label = tk.Label(root, text="Destination folder you want to copy the files/subfolders to:")
copy_to_label.grid(row=8, column=0)
copy_to_entry = tk.Entry(root, width=120, textvariable=copy_to_folder)
copy_to_entry.grid(row=8, column=1)
copy_to_button = tk.Button(root, text="Browse", command=select_copy_to_folder)
copy_to_button.grid(row=8, column=2)

empty = tk.Label(root, text="")
empty.grid(row=9, column=1)
empty = tk.Label(root, text="")
empty.grid(row=10, column=1)

recurisively_check = tk.IntVar()
recurisively_check_label = tk.Label(root, text="Copy files/subfolders recursively?")
recurisively_check_label.grid(row=11, column=0)
recurisively_check_checkbutton = tk.Checkbutton(root, text="", variable=recurisively_check)
recurisively_check_checkbutton.grid(row=11, column=1)
recurisively_check_note = tk.Label(root, text='If unchecked, files/subfolders will be copied from the original folder only.')
recurisively_check_note.grid(row=12, column=1)
recurisively_check_note = tk.Label(root, text='If checked, files/subfolders will be copied from the original folder and its subfolders, sub-subfolder, ... recursively.')
recurisively_check_note.grid(row=13, column=1)

empty = tk.Label(root, text="")
empty.grid(row=14, column=1)
empty = tk.Label(root, text="")
empty.grid(row=15, column=1)

run_button = tk.Button(root, text="Copy data files/subfolders", command=run_code)
run_button.grid(row=16, column=1)

root.mainloop()